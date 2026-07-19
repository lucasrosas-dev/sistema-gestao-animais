from __future__ import annotations

import csv
import io
import zipfile
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable

from fastapi.responses import Response, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from ..utils.formatting import csv_safe, safe_filename


def csv_response(filename: str, headers: list[str], rows: Iterable[Iterable[Any]]) -> Response:
    output = io.StringIO(newline="")
    writer = csv.writer(output, delimiter=";", lineterminator="\n")
    writer.writerow(headers)
    for row in rows:
        writer.writerow([csv_safe(value) for value in row])
    payload = "\ufeff" + output.getvalue()
    return Response(
        content=payload.encode("utf-8"),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{safe_filename(filename)}.csv"'},
    )


def xlsx_response(
    filename: str,
    title: str,
    headers: list[str],
    rows: list[list[Any]],
    *,
    generated_by: str,
    period_label: str,
    filters_label: str = "",
    sheet_name: str = "Dados",
    summary: list[tuple[str, Any]] | None = None,
) -> StreamingResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    ws["A1"] = title
    ws["A1"].font = Font(size=16, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(headers)))
    ws["A2"] = f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} por {generated_by}"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(1, len(headers)))
    ws["A3"] = f"Período: {period_label}" + (f" | Filtros: {filters_label}" if filters_label else "")
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max(1, len(headers)))
    header_row = 5
    fill = PatternFill("solid", fgColor="285D48")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(header_row, col, header)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
    for r_index, row in enumerate(rows, header_row + 1):
        for c_index, value in enumerate(row, 1):
            cell = ws.cell(r_index, c_index, value)
            if isinstance(value, Decimal):
                cell.value = float(value)
                cell.number_format = '#,##0.00'
            elif hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
                cell.number_format = "dd/mm/yyyy"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{max(header_row, header_row + len(rows))}"
    ws.freeze_panes = f"A{header_row + 1}"
    for index, header in enumerate(headers, 1):
        max_len = len(str(header))
        for row in ws.iter_rows(min_row=header_row + 1, max_row=min(header_row + len(rows), header_row + 500), min_col=index, max_col=index):
            max_len = max(max_len, len(str(row[0].value or "")))
        ws.column_dimensions[get_column_letter(index)].width = min(max_len + 3, 45)

    if summary:
        sws = wb.create_sheet("Resumo")
        sws["A1"] = title
        sws["A1"].font = Font(size=16, bold=True)
        sws.append([])
        sws.append(["Indicador", "Valor"])
        sws["A3"].font = sws["B3"].font = Font(bold=True, color="FFFFFF")
        sws["A3"].fill = sws["B3"].fill = fill
        for label, value in summary:
            sws.append([label, float(value) if isinstance(value, Decimal) else value])
        sws.column_dimensions["A"].width = 38
        sws.column_dimensions["B"].width = 22

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_filename(filename)}.xlsx"'},
    )


def pdf_response(
    filename: str,
    title: str,
    headers: list[str],
    rows: list[list[Any]],
    *,
    generated_by: str,
    period_label: str,
    landscape_mode: bool = False,
    summary: list[tuple[str, Any]] | None = None,
) -> StreamingResponse:
    stream = io.BytesIO()
    pagesize = landscape(A4) if landscape_mode else A4

    def footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.drawString(15 * mm, 10 * mm, "Sistema de Gestão de Animais")
        canvas.drawRightString(pagesize[0] - 15 * mm, 10 * mm, f"Página {doc.page}")
        canvas.restoreState()

    doc = SimpleDocTemplate(stream, pagesize=pagesize, leftMargin=12 * mm, rightMargin=12 * mm, topMargin=14 * mm, bottomMargin=16 * mm)
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="CenterSmall", parent=styles["Normal"], alignment=TA_CENTER, fontSize=8, leading=10))
    story = [Paragraph(title, styles["Title"]), Spacer(1, 4), Paragraph(f"Período: {period_label}", styles["Normal"]), Paragraph(f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} por {generated_by}", styles["Normal"]), Spacer(1, 10)]
    if summary:
        summary_data = [[Paragraph("Indicador", styles["CenterSmall"]), Paragraph("Valor", styles["CenterSmall"])]] + [[str(a), str(b)] for a, b in summary]
        summary_table = Table(summary_data, colWidths=[65 * mm, 45 * mm])
        summary_table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#285D48")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("GRID", (0, 0), (-1, -1), .35, colors.HexColor("#dfe5ea")), ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTSIZE", (0, 0), (-1, -1), 8), ("VALIGN", (0, 0), (-1, -1), "MIDDLE")]))
        story.extend([summary_table, Spacer(1, 12)])
    data = [[Paragraph(str(item), styles["CenterSmall"]) for item in headers]]
    for row in rows:
        data.append([Paragraph(str(item if item is not None else ""), styles["CenterSmall"]) for item in row])
    available_width = pagesize[0] - 24 * mm
    col_width = available_width / max(1, len(headers))
    table = Table(data, colWidths=[col_width] * len(headers), repeatRows=1)
    table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#285D48")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("GRID", (0, 0), (-1, -1), .3, colors.HexColor("#cbd4db")), ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafb")]), ("VALIGN", (0, 0), (-1, -1), "TOP"), ("LEFTPADDING", (0, 0), (-1, -1), 3), ("RIGHTPADDING", (0, 0), (-1, -1), 3), ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4)]))
    story.append(table)
    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    stream.seek(0)
    return StreamingResponse(stream, media_type="application/pdf", headers={"Content-Disposition": f'attachment; filename="{safe_filename(filename)}.pdf"'})


def zip_response(filename: str, files: dict[str, bytes]) -> StreamingResponse:
    stream = io.BytesIO()
    with zipfile.ZipFile(stream, "w", zipfile.ZIP_DEFLATED) as archive:
        for name, payload in files.items():
            source = Path(name)
            stem = safe_filename(source.stem or 'dados', 'dados')
            archive.writestr(stem + source.suffix.lower(), payload)
    stream.seek(0)
    return StreamingResponse(stream, media_type="application/zip", headers={"Content-Disposition": f'attachment; filename="{safe_filename(filename)}.zip"'})
